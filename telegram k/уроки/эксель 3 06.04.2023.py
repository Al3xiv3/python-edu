header = ['name', 'surname', 'age']
ppl = [
    ['Alex', 'Siv', '25'],
    ['Ilya', 'Zhen', '18'],
    ]

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['B1'] = header[0]
ws['C1'] = header[1]
ws['D1'] = header[2]

ind = 2

for people in ppl:
    ws["A" + str(ind)] = ind - 1
    ws["B" + str(ind)] = people[0]
    ws["C" + str(ind)] = people[1]
    ws["D" + str(ind)] = int(people[2])
    ind += 1


wb.save('ppl_sheet.xlsx')

print(ws['C2:C3'][0][0].value)
