from openpyxl import Workbook

wb = Workbook()

ws1 = wb.create_sheet("Mysheet")
ws = wb["Mysheet"]

let = 'ABCDE'
index = let.index('A') + 1


for letter in let:
    for num in range(1, 6):
        ws[letter + str(num)] = letter*num


wb.save('test.xlsx')