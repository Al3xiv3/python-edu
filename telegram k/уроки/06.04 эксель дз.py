header = ['id', 'name', 'grade']
game = [
    ['1', 'Atomic Heart', '4.00'],
    ['2', 'Far Cry', '4.40'],
    ['3', 'Baldi', '5.00'],
    ]

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = header[0]
ws['B1'] = header[1]
ws['C1'] = header[2]

ind = 2

for games in game:
    ws["A" + str(ind)] = int(games[0])
    ws["B" + str(ind)] = games[1]
    ws["C" + str(ind)] = games[2]
    ind += 1

wb.save('GamesHomeTask.xlsx')

