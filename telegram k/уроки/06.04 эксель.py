from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet("Mysheet")

let = 'ABCDE'
index = let.index('A') + 1


for letter in let:
    for num in range(1, 6):
        ws[letter + str(num)] = (let.index(letter) + 1)*num


wb.save('test.xlsx')
